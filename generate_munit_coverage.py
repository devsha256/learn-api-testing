"""
MUnit Consolidated Coverage Report
===================================
Reads from every project listed in a projects file under ROOT_FOLDER:
  - target/site/munit/coverage/munit-coverage.json   (Mule 4 coverage)
  - target/surefire-reports/*.xml                     (test pass/fail)

Produces an Excel workbook with two sheets:
  1. application_coverage  — one row per application
  2. flow_coverage         — one row per flow, application as foreign key

Usage:
    python munit_report.py
    python munit_report.py --root C:/Users/Saddam/Projects --projects C:/Users/Saddam/Projects/projects.csv --out C:/reports/MUnit_Report.xlsx
"""

import argparse
import json
import os
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── defaults ──────────────────────────────────────────────────────────────────
DEFAULT_ROOT     = "C:/Users/Saddam/Projects"
DEFAULT_PROJECTS = "C:/Users/Saddam/Projects/projects.csv"   # one project name per line
DEFAULT_OUT      = "MUnit_Coverage_Report.xlsx"

# ── colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG = "1F3864"
CLR_HEADER_FG = "FFFFFF"
CLR_SUBHDR_BG = "2E75B6"
CLR_SUBHDR_FG = "FFFFFF"
CLR_ROW_ODD   = "EBF3FB"
CLR_ROW_EVEN  = "FFFFFF"
CLR_GREEN_BG  = "E2EFDA"
CLR_AMBER_BG  = "FFF2CC"
CLR_RED_BG    = "FCE4D6"
CLR_BORDER    = "BDD7EE"

THIN   = Side(style="thin", color=CLR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── helpers ───────────────────────────────────────────────────────────────────

def coverage_fill(pct: float) -> PatternFill:
    if pct >= 80:
        return PatternFill("solid", fgColor=CLR_GREEN_BG)
    if pct >= 50:
        return PatternFill("solid", fgColor=CLR_AMBER_BG)
    return PatternFill("solid", fgColor=CLR_RED_BG)


def coverage_label(pct: float) -> str:
    if pct >= 80:
        return "GOOD"
    if pct >= 50:
        return "PARTIAL"
    return "LOW"


def header_cell(ws, row, col, value, bold=True, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c


def data_cell(ws, row, col, value, bold=False, fill=None, number_format=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    return c


def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def row_fill(row_idx: int) -> PatternFill:
    colour = CLR_ROW_ODD if row_idx % 2 == 1 else CLR_ROW_EVEN
    return PatternFill("solid", fgColor=colour)


# ── data collection ───────────────────────────────────────────────────────────

def load_project_list(projects_file: Path) -> list[str]:
    """
    Reads a plain text / CSV file with one project folder name per line.
    Blank lines and lines starting with # are ignored.
    The file is the same projects.csv already used by the PowerShell script.
    """
    if not projects_file.exists():
        raise FileNotFoundError(
            f"Projects file not found: {projects_file}\n"
            f"Create it with one project folder name per line, e.g.:\n"
            f"  customer-experience-api\n"
            f"  order-process-api\n"
        )
    names = []
    for line in projects_file.read_text(encoding="utf-8").splitlines():
        name = line.strip()
        if name and not name.startswith("#"):
            names.append(name)
    if not names:
        raise ValueError(f"Projects file is empty or has no valid entries: {projects_file}")
    return names


def resolve_projects(root: Path, project_names: list[str]) -> list[Path]:
    """
    Resolves project names to full paths under root.
    Warns and skips any that do not exist or have no pom.xml.
    """
    resolved = []
    for name in project_names:
        path = root / name
        if not path.exists():
            print(f"  [!] SKIP: '{name}' — directory not found at {path}")
            continue
        if not (path / "pom.xml").exists():
            print(f"  [!] SKIP: '{name}' — no pom.xml found at {path}")
            continue
        resolved.append(path)
    return resolved


def load_coverage_json(project: Path) -> dict | None:
    candidates = [
        project / "target" / "site" / "munit" / "coverage" / "munit-coverage.json",
        project / "target" / "munit-reports" / "coverage-json" / "report.json",
    ]
    for path in candidates:
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                return None
    return None


def load_surefire_reports(project: Path) -> dict:
    surefire_dir = project / "target" / "surefire-reports"
    result = {"tests": 0, "failures": 0, "errors": 0, "skipped": 0, "test_cases": []}
    if not surefire_dir.exists():
        return result
    for xml_file in surefire_dir.glob("*.xml"):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            suites = [root] if root.tag == "testsuite" else root.findall("testsuite")
            for suite in suites:
                result["tests"]    += int(suite.get("tests",    0))
                result["failures"] += int(suite.get("failures", 0))
                result["errors"]   += int(suite.get("errors",   0))
                result["skipped"]  += int(suite.get("skipped",  0))
                for tc in suite.findall("testcase"):
                    failure = tc.find("failure")
                    error   = tc.find("error")
                    skipped = tc.find("skipped")
                    if failure is not None:
                        status, message = "FAILED",  failure.get("message", "")[:200]
                    elif error is not None:
                        status, message = "ERROR",   error.get("message", "")[:200]
                    elif skipped is not None:
                        status, message = "SKIPPED", ""
                    else:
                        status, message = "PASSED",  ""
                    result["test_cases"].append({
                        "name":      tc.get("name", ""),
                        "classname": tc.get("classname", ""),
                        "status":    status,
                        "time":      round(float(tc.get("time", 0)), 3),
                        "message":   message,
                    })
        except Exception:
            continue
    return result


def collect_all(root: Path, project_names: list[str]) -> list[dict]:
    """
    Resolves the supplied project names under root, loads coverage + surefire
    data for each, and returns a list of app dicts in the same order as the
    projects file.
    """
    projects = resolve_projects(root, project_names)
    apps = []
    for project in projects:
        coverage = load_coverage_json(project)
        surefire = load_surefire_reports(project)

        app_cov    = coverage.get("coverage",               0.0) if coverage else 0.0
        flow_count = coverage.get("flowCount",              0)   if coverage else 0
        proc_count = coverage.get("processorCount",         0)   if coverage else 0
        cov_procs  = coverage.get("coveredProcessorCount",  0)   if coverage else 0
        files      = coverage.get("files",                  [])  if coverage else []

        apps.append({
            "name":                    project.name,
            "path":                    str(project),
            "coverage_pct":            round(app_cov, 2),
            "flow_count":              flow_count,
            "processor_count":         proc_count,
            "covered_processor_count": cov_procs,
            "resource_count":          len(files),
            "tests":                   surefire["tests"],
            "failures":                surefire["failures"],
            "errors":                  surefire["errors"],
            "skipped":                 surefire["skipped"],
            "passed":                  surefire["tests"] - surefire["failures"] - surefire["errors"] - surefire["skipped"],
            "test_cases":              surefire["test_cases"],
            "files":                   files,
            "coverage_found":          coverage is not None,
        })
    return apps


# ── sheet builders ────────────────────────────────────────────────────────────

def build_application_coverage(wb: Workbook, apps: list[dict]):
    ws = wb.create_sheet("application_coverage")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:O1")
    title = ws["A1"]
    title.value     = f"MUnit Application Coverage Report  •  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_FG)
    title.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "#", "Application", "Coverage %", "Status",
        "Resources (XML)", "Total Flows", "Total Processors",
        "Covered Processors", "Uncovered Processors",
        "Total Tests", "Passed", "Failed", "Errors", "Skipped", "Pass Rate %",
    ]
    for col, h in enumerate(headers, start=1):
        header_cell(ws, 2, col, h, bg=CLR_SUBHDR_BG)
    ws.row_dimensions[2].height = 32

    for idx, app in enumerate(apps, start=1):
        r         = idx + 2
        rf        = row_fill(idx)
        cvf       = coverage_fill(app["coverage_pct"])
        uncovered = app["processor_count"] - app["covered_processor_count"]
        pass_rate = round(app["passed"] / app["tests"] * 100, 2) if app["tests"] > 0 else 0.0

        data_cell(ws, r,  1, idx,                             fill=rf,  align="center")
        data_cell(ws, r,  2, app["name"],                     fill=rf,  bold=True)
        data_cell(ws, r,  3, app["coverage_pct"],             fill=cvf, align="center", number_format='0.00"%"')
        data_cell(ws, r,  4, coverage_label(app["coverage_pct"]), fill=cvf, align="center", bold=True)
        data_cell(ws, r,  5, app["resource_count"],           fill=rf,  align="center")
        data_cell(ws, r,  6, app["flow_count"],               fill=rf,  align="center")
        data_cell(ws, r,  7, app["processor_count"],          fill=rf,  align="center")
        data_cell(ws, r,  8, app["covered_processor_count"],  fill=rf,  align="center")
        data_cell(ws, r,  9, uncovered,                       fill=rf,  align="center")
        data_cell(ws, r, 10, app["tests"],                    fill=rf,  align="center")
        data_cell(ws, r, 11, app["passed"],                   fill=rf,  align="center")
        data_cell(ws, r, 12, app["failures"],
                  fill=PatternFill("solid", fgColor=CLR_RED_BG) if app["failures"] > 0 else rf,
                  align="center", bold=app["failures"] > 0)
        data_cell(ws, r, 13, app["errors"],
                  fill=PatternFill("solid", fgColor=CLR_AMBER_BG) if app["errors"] > 0 else rf,
                  align="center", bold=app["errors"] > 0)
        data_cell(ws, r, 14, app["skipped"],                  fill=rf,  align="center")
        data_cell(ws, r, 15, pass_rate,
                  fill=coverage_fill(pass_rate), align="center", number_format='0.00"%"')
        ws.row_dimensions[r].height = 18

    last_data = len(apps) + 2
    total_row = last_data + 1
    tf = PatternFill("solid", fgColor="1F3864")

    def tot(col, formula):
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = tf
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER

    ws.merge_cells(f"A{total_row}:B{total_row}")
    t = ws.cell(row=total_row, column=1, value="TOTALS / AVERAGES")
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    t.fill      = tf
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.border    = BORDER

    first, last = 3, last_data
    tot(3,  f"=AVERAGE(C{first}:C{last})"); ws.cell(row=total_row, column=3).number_format = '0.00"%"'
    tot(4,  "")
    tot(5,  f"=SUM(E{first}:E{last})")
    tot(6,  f"=SUM(F{first}:F{last})")
    tot(7,  f"=SUM(G{first}:G{last})")
    tot(8,  f"=SUM(H{first}:H{last})")
    tot(9,  f"=SUM(I{first}:I{last})")
    tot(10, f"=SUM(J{first}:J{last})")
    tot(11, f"=SUM(K{first}:K{last})")
    tot(12, f"=SUM(L{first}:L{last})")
    tot(13, f"=SUM(M{first}:M{last})")
    tot(14, f"=SUM(N{first}:N{last})")
    tot(15, f"=IFERROR(K{total_row}/J{total_row}*100,0)"); ws.cell(row=total_row, column=15).number_format = '0.00"%"'
    ws.row_dimensions[total_row].height = 22

    set_col_widths(ws, [5, 36, 14, 12, 16, 14, 18, 20, 22, 14, 10, 10, 10, 10, 14])


def build_flow_coverage(wb: Workbook, apps: list[dict]):
    ws = wb.create_sheet("flow_coverage")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value     = f"MUnit Flow Coverage Report  •  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_FG)
    title.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "#", "Application", "Resource File",
        "Resource Coverage %", "Resource Weight %",
        "Flow Name", "Flow Type", "Flow Coverage %",
        "Coverage Status", "Total Processors", "Covered Processors",
    ]
    for col, h in enumerate(headers, start=1):
        header_cell(ws, 2, col, h, bg=CLR_SUBHDR_BG)
    ws.row_dimensions[2].height = 32

    row_num    = 3
    global_idx = 1

    for app in apps:
        if not app["files"]:
            rf = row_fill(global_idx)
            data_cell(ws, row_num,  1, global_idx,               fill=rf, align="center")
            data_cell(ws, row_num,  2, app["name"],              fill=rf, bold=True)
            data_cell(ws, row_num,  3, "N/A",                    fill=rf)
            data_cell(ws, row_num,  4, "N/A",                    fill=rf, align="center")
            data_cell(ws, row_num,  5, "N/A",                    fill=rf, align="center")
            data_cell(ws, row_num,  6, "No coverage report found", fill=rf)
            data_cell(ws, row_num,  7, "N/A",                    fill=rf, align="center")
            data_cell(ws, row_num,  8, "N/A",                    fill=rf, align="center")
            data_cell(ws, row_num,  9, "NO REPORT",
                      fill=PatternFill("solid", fgColor=CLR_AMBER_BG), align="center", bold=True)
            data_cell(ws, row_num, 10, "N/A",                    fill=rf, align="center")
            data_cell(ws, row_num, 11, "N/A",                    fill=rf, align="center")
            ws.row_dimensions[row_num].height = 18
            row_num    += 1
            global_idx += 1
            continue

        for file_entry in app["files"]:
            file_name   = file_entry.get("name", "")
            file_cov    = round(float(file_entry.get("coverage", 0.0)), 2)
            file_weight = round(float(file_entry.get("weight",   0.0)), 2)
            flows       = file_entry.get("flows", [])

            if not flows:
                rf = row_fill(global_idx)
                data_cell(ws, row_num,  1, global_idx,  fill=rf, align="center")
                data_cell(ws, row_num,  2, app["name"], fill=rf, bold=True)
                data_cell(ws, row_num,  3, file_name,   fill=rf)
                data_cell(ws, row_num,  4, file_cov,    fill=coverage_fill(file_cov), align="center", number_format='0.00"%"')
                data_cell(ws, row_num,  5, file_weight, fill=rf, align="center", number_format='0.00"%"')
                data_cell(ws, row_num,  6, "(no flows)", fill=rf)
                for c in range(7, 12):
                    data_cell(ws, row_num, c, "", fill=rf, align="center")
                ws.row_dimensions[row_num].height = 18
                row_num    += 1
                global_idx += 1
                continue

            for flow in flows:
                rf         = row_fill(global_idx)
                flow_cov   = round(float(flow.get("coverage", 0.0)), 2)
                cvf        = coverage_fill(flow_cov)

                data_cell(ws, row_num,  1, global_idx,                    fill=rf,  align="center")
                data_cell(ws, row_num,  2, app["name"],                   fill=rf,  bold=True)
                data_cell(ws, row_num,  3, file_name,                     fill=rf)
                data_cell(ws, row_num,  4, file_cov,                      fill=coverage_fill(file_cov), align="center", number_format='0.00"%"')
                data_cell(ws, row_num,  5, file_weight,                   fill=rf,  align="center", number_format='0.00"%"')
                data_cell(ws, row_num,  6, flow.get("name", ""),          fill=rf,  bold=True)
                data_cell(ws, row_num,  7, flow.get("type", ""),          fill=rf,  align="center")
                data_cell(ws, row_num,  8, flow_cov,                      fill=cvf, align="center", number_format='0.00"%"')
                data_cell(ws, row_num,  9, coverage_label(flow_cov),      fill=cvf, align="center", bold=True)
                data_cell(ws, row_num, 10, flow.get("messageProcessorCount", 0),  fill=rf, align="center")
                data_cell(ws, row_num, 11, flow.get("coveredProcessorCount",  0), fill=rf, align="center")
                ws.row_dimensions[row_num].height = 18
                row_num    += 1
                global_idx += 1

    set_col_widths(ws, [5, 36, 32, 18, 18, 42, 14, 16, 16, 18, 20])


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="MUnit Consolidated Coverage Report")
    parser.add_argument(
        "--root",
        default=DEFAULT_ROOT,
        help=f"Root folder containing Mule projects (default: {DEFAULT_ROOT})"
    )
    parser.add_argument(
        "--projects",
        default=DEFAULT_PROJECTS,
        help=f"Path to file listing project folder names, one per line (default: {DEFAULT_PROJECTS})"
    )
    parser.add_argument(
        "--out",
        default=DEFAULT_OUT,
        help=f"Output Excel file path (default: {DEFAULT_OUT})"
    )
    args = parser.parse_args()

    root          = Path(args.root)
    projects_file = Path(args.projects)
    out_path      = Path(args.out)

    if not root.exists():
        print(f"[ERROR] Root folder not found: {root}")
        return

    print(f"[*] Root folder   : {root}")
    print(f"[*] Projects file : {projects_file}")

    try:
        project_names = load_project_list(projects_file)
    except (FileNotFoundError, ValueError) as e:
        print(f"[ERROR] {e}")
        return

    print(f"[*] Projects requested : {len(project_names)}")
    for n in project_names:
        print(f"      - {n}")

    apps = collect_all(root, project_names)

    if not apps:
        print("[!] No valid projects resolved. Nothing to report.")
        return

    print(f"\n[*] Building report for {len(apps)} resolved projects...")

    wb = Workbook()
    wb.remove(wb.active)

    build_application_coverage(wb, apps)
    build_flow_coverage(wb, apps)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    total_tests    = sum(a["tests"]    for a in apps)
    total_failures = sum(a["failures"] for a in apps)
    total_errors   = sum(a["errors"]   for a in apps)
    avg_coverage   = round(sum(a["coverage_pct"] for a in apps) / len(apps), 2)

    print()
    print("=" * 60)
    print("  MUNIT COVERAGE REPORT SUMMARY")
    print("=" * 60)
    print(f"  Projects requested   : {len(project_names)}")
    print(f"  Projects resolved    : {len(apps)}")
    print(f"  Avg application cov  : {avg_coverage}%")
    print(f"  GOOD  (>=80%)        : {sum(1 for a in apps if a['coverage_pct'] >= 80)}")
    print(f"  PARTIAL (50-79%)     : {sum(1 for a in apps if 50 <= a['coverage_pct'] < 80)}")
    print(f"  LOW   (<50%)         : {sum(1 for a in apps if a['coverage_pct'] < 50)}")
    print(f"  No coverage report   : {sum(1 for a in apps if not a['coverage_found'])}")
    print(f"  Total tests          : {total_tests}")
    print(f"  Failures + Errors    : {total_failures + total_errors}")
    print("=" * 60)
    print(f"[*] Report saved to: {out_path.resolve()}")


if __name__ == "__main__":
    main()
